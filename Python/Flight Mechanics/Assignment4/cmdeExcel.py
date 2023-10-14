import openpyxl
import ExcelSorter as ExS
import FinalFunc


data0 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_0.xlsx')
data1 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_4.xlsx')
data2 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_7.xlsx')
data3 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus4.xlsx')
data4 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus7.xlsx')
data5 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus14.xlsx')

note0 = data0.active
note1 = data1.active
note2 = data2.active
note3 = data3.active
note4 = data4.active
note5 = data5.active

notes = [note0, note1, note2, note3, note4, note5]

print("[AoA = 0]\n")
FinalFunc.Calculation(notes)


data0 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_8_Elev_Varying/alpha_8_elev_0.xlsx')
data1 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_8_Elev_Varying/alpha_8_elev_4.xlsx')
data2 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_8_Elev_Varying/alpha_8_elev_7.xlsx')
data3 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_8_Elev_Varying/alpha_8_elev_minus4.xlsx')
data4 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_8_Elev_Varying/alpha_8_elev_minus7.xlsx')
data5 = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_8_Elev_Varying/alpha_8_elev_minus14.xlsx')

note0 = data0.active
note1 = data1.active
note2 = data2.active
note3 = data3.active
note4 = data4.active
note5 = data5.active

notes = [note0, note1, note2, note3, note4, note5]
print("[AoA = 8]\n")
FinalFunc.Calculation(notes)






