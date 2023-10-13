import openpyxl
import ExcelSorter as ExS

data0 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_0.xlsx')
data1 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_4.xlsx')
data2 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_7.xlsx')
data3 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus4.xlsx')
data4 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus7.xlsx')
data5 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus14.xlsx')

note0 = data0.active
note1 = data1.active
note2 = data2.active
note3 = data3.active
note4 = data4.active
note5 = data5.active

notes = [note0, note1, note2, note3, note4, note5]


i = 1
for count, value in enumerate(notes):

    x = ExS.numberSorter(value)
    print("Sheet " + str(count+1) + "\nDrag: [" + str(x[0]) + " N]\nPitching Moment: [" + str(x[1]) + " N]\nDrag reducted from Sting: [" + str(x[2]) + " N]")
    print('Total number of rows: '+str(value.max_row)+'. And total number of columns: '+str(value.max_column) +"\n")
    i+=1







