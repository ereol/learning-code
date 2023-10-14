def numberSorter(note):
    import math
    import openpyxl
    # print('Total number of rows: '+str(note.max_row)+'. And total number of columns: '+str(note.max_column))
    data = openpyxl.load_workbook(r'D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_35.xlsx')
    ws = data.active

    i = 1
    drag = 0
    pitchm = 0
    sting = 0
    dragsting=0

    for i in range(1,note.max_row):
        drag+=(note.cell(row=i,column=2).value)
        sting+=(ws.cell(row=i,column=2).value)
        dragsting+=((note.cell(row=i,column=2).value)-(ws.cell(row=i,column=2).value))
        pitchm+=(note.cell(row=i,column=4).value)
        i+=1

    
    pitchm /= note.max_row
    drag /= note.max_row
    dragsting/=note.max_row
    sting /=ws.max_row
    return(drag,pitchm, dragsting, sting)