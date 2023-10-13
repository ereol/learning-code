
import openpyxl
import ExcelSorter as ExS

data0 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_10.xlsx')
data1 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_15.xlsx')
data2 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_20.xlsx')
data3 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_25.xlsx')
data4 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_30.xlsx')
data5 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_35.xlsx')
data6 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Sting_Only/sting_speed_40.xlsx')

note0 = data0.active
note1 = data1.active
note2 = data2.active
note3 = data3.active
note4 = data4.active
note5 = data5.active
note6 = data6.active

notes = [note0, note1, note2, note3, note4, note5, note6]

for count, value in enumerate(notes):

    x = ExS.numberSorter(value)
    print("Sheet " + str(count+1) + ": [" + str(x[1]) + "]")
    print('Total number of rows: '+str(value.max_row)+'. And total number of columns: '+str(value.max_column) +"\n")