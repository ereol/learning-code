import openpyxl
import ExcelSorter as ExS

data0 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_0.xlsx')
data1 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_4.xlsx')
data2 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_7.xlsx')
data3 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus4.xlsx')
data4 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus7.xlsx')
data5 = openpyxl.load_workbook('D:/Desktop/Academic/2023/Semester 2/AERO - Flight Dynamics/Alpha_0_Elev_Varying/alpha_0_elev_minus14.xlsx')

note0 = data0.active
note1 = data1.active
note2 = data2.active
note3 = data3.active
note4 = data4.active
note5 = data5.active

notes = [ note0, note1, note2, note3, note4, note5]


i = 1
for value in (notes):
    x = ExS.numberSorter(value)
    print("Sheet" + str(i))
    print(x)
    i+=1






# print('Total number of rows: '+str(note.max_row)+'. And total number of columns: '+str(note.max_column))

# i = 1
# f = 1
# x = 0

# for f in range(1,note.max_column):
#     for i in range(1,note.max_row):
#         x+=note.cell(row=i,column=f).value

#         i+=1
#     f+=1

# x /= (note.max_row * note.max_column)
# print(x)






